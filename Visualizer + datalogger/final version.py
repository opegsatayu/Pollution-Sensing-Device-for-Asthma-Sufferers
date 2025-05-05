import serial, time
import pandas as pd
import matplotlib.pyplot as plt  #for graph
from matplotlib.animation import FuncAnimation #for animation + update 
from matplotlib.patches import Circle #for shapes(circles)
from openpyxl import load_workbook
import matplotlib.dates as mdates  # for time 

#python will read what arduino code print out in order 
PORT = 'COM3'
BAUD_RATE = 115200
ser = serial.Serial(PORT, BAUD_RATE, timeout=1)
excel_file = r'C:\Users\opeg\OneDrive - University of Dundee\year 4\sem 1\Project\code\Python test VS code\Final python\datalog.xlsx'

data_log = []
log_time = time.time() #log 1 data per second 


# CURRENT SET UP FOR VIVA, 3 SECOND OVER THRESHOLD OF 400 = GREEN LED 
# 6 total SECONDS FOR RED TO COME 
# 2 SECOND = SAFE
# 3 SECOND  FOR RE-EXPOSE  

#----------------------circles set up----------------------------------------------------------------------

#plot setip 
fig, ax = plt.subplots()
ax.set_aspect('equal') # make the axis scaling the same 
ax.set_xlim(-1.2, 1.2 ) #set x-axis limit =1.2 to 2 units
ax.set_ylim(-1.2, 1.2) #set y-axis limit =1.2 to 2 units
ax.axis('off') #dont show axis 

#ozone
outer_circle_ozone = Circle((0, 0.8), 0.35, edgecolor='black', fill=False, linewidth=3)
inner_circle_ozone = Circle((0, 0.8), 0.1, color='lightblue')
ax.add_patch(outer_circle_ozone)
ax.add_patch(inner_circle_ozone)
text_ozone = ax.text(0, 0.8, "", fontsize=10, ha='center', va='center') #using ax.text to select the axis of where the text should be (same as ozone circle) 

#NO2
outer_circle_no2 = Circle((-0.8, 0), 0.35, edgecolor='black', fill=False, linewidth=3)
inner_circle_no2 = Circle((-0.8, 0), 0.1, color='lightgreen')
ax.add_patch(outer_circle_no2)
ax.add_patch(inner_circle_no2)
text_NO2 = ax.text(-0.8, 0, "", fontsize=10, ha='center', va='center')

#PM2.5 
outer_circle_pm25 = Circle((-0.5, -0.8), 0.35, edgecolor='black', fill=False, linewidth=3)
inner_circle_pm25 = Circle((-0.5, -0.8), 0.1, color='lightcoral')
ax.add_patch(outer_circle_pm25)
ax.add_patch(inner_circle_pm25)
text_PM25 = ax.text(-0.5, -0.8, "", fontsize=10, ha='center', va='center')

#PM10
outer_circle_pm10 = Circle((0.5, -0.8), 0.35, edgecolor='black', fill=False, linewidth=3)
inner_circle_pm10 = Circle((0.5, -0.8), 0.1, color='pink')
ax.add_patch(outer_circle_pm10)
ax.add_patch(inner_circle_pm10)
text_PM10 = ax.text(0.5, -0.8, "", fontsize=10, ha='center', va='center')

#temp
outer_circle_temp = Circle((0.8, 0), 0.35, edgecolor='black', fill=False, linewidth=3)
inner_circle_temp = Circle((0.8, 0), 0.1, color='lightyellow')
ax.add_patch(outer_circle_temp)
ax.add_patch(inner_circle_temp)
text_temp = ax.text(0.8, 0, "", fontsize=10, ha='center', va='center')

#----------------------- realtime updating ------------------------------------
 
def update(frame):
    global log_time
    if ser.in_waiting > 0:
        try:
            line = ser.readline().decode('utf-8')

            Ozone_reading, NO2_reading, PM25_reading, PM10_reading, temp_reading = line.split(',')
            Ozone_reading, NO2_reading, PM25_reading, PM10_reading, temp_reading = float(Ozone_reading), float(NO2_reading), float(PM25_reading), float(PM10_reading), float(temp_reading)

            #ozone 
            inner_circle_ozone.set_radius(min(Ozone_reading / 0.2 * 0.35, 0.35)) # was 0.2 ozone reading / threshold value (if ozone reading reaches the same as thresold, circle will be filled)
            text_ozone.set_text(f"{Ozone_reading:.3f} PPM") #3 decimal 
            
            #NO2
            inner_circle_no2.set_radius(min(NO2_reading / 1 * 0.35, 0.35)) #threshold of 1 ppm is bad 
            text_NO2.set_text(f"{NO2_reading:.3f} PPM")

            #pm2.5
            inner_circle_pm25.set_radius(min(PM25_reading / 50 * 0.35, 0.35)) #threshold of 50 ug/m3 is bad 
            text_PM25.set_text(f"{PM25_reading:.3f} µg/m3")
            #pm10
            inner_circle_pm10.set_radius(min(PM10_reading / 150 * 0.35, 0.35)) #threshold of 150 ug/m3 is bad
            text_PM10.set_text(f"{PM10_reading:.3f} µg/m3")
            #temp
            inner_circle_temp.set_radius(min(temp_reading / 50 * 0.35, 0.35)) #threshold of 35 celcius 
            text_temp.set_text(f"{temp_reading:.3f} °C")


            fig.suptitle("Ozone (lightblue), NO2 (lightgreen), PM2.5 (lightcoral), PM10 (pink), Temperature (lightyellow)", fontsize=9)

                         
       
            now = time.time() #time right now
            if now - log_time >= 5.0: #if 5 second passed, update the excel
                timestamp = time.strftime('%H:%M:%S') #day/month/year and hour:min:second format 
                data_log.append({'Time': timestamp, 'Ozone': Ozone_reading, 'NO2': NO2_reading, 'PM2.5': PM25_reading, 'PM10': PM10_reading, 'temp.': temp_reading})
                log_time = now
        except:
            pass

ani = FuncAnimation(fig, update, interval=50, cache_frame_data=False)
print("close the window to stop recording")
plt.show()
ser.close()

if data_log:
        df = pd.DataFrame(data_log)
        df.to_excel(excel_file, sheet_name='Sheet1', index=False)
        print(f"Saved {len(data_log)} rows to {excel_file}")
else:
        print("no data to save.")

#--------------------------averaging for excel------------------------


wb = load_workbook(excel_file)
ws = wb['Sheet1']

#avg colum
ws['G1'] = "Average Ozone (PPM)"
ws['H1'] = "Average NO2 (PPM)"
ws['I1'] = "Average PM2.5 (µg/m³)"
ws['J1'] = "Average PM10 (µg/m³)"
ws['K1'] = "Average Temperature (°C)"

#average formula in excel 
ws['G2'] = "=AVERAGE(B:B)"
ws['H2'] = "=AVERAGE(C:C)"
ws['I2'] = "=AVERAGE(D:D)"
ws['J2'] = "=AVERAGE(E:E)"
ws['K2'] = "=AVERAGE(F:F)"

wb.save(excel_file)

# --------------- Plotting Sensor Readings vs Time  ------------------------
df_plot = pd.read_excel(excel_file, sheet_name='Sheet1')
df_plot_time = pd.to_datetime(df_plot['Time'], format='%H:%M:%S') # Convert 'Time' to a datetime using the HH:MM:SS format

# Create a new figure
fig, ax = plt.subplots(figsize=(12, 6))


# plot each sensor reading vs the datetime 'Time' column
plt.plot(df_plot_time, df_plot['Ozone'], marker='', label='Ozone (PPM)')
plt.plot(df_plot_time, df_plot['NO2'], marker='', label='NO2 (PPM)')
plt.plot(df_plot_time, df_plot['PM2.5'], marker='', label='PM2.5 (µg/m³)')
plt.plot(df_plot_time, df_plot['PM10'], marker='', label='PM10 (µg/m³)')
plt.plot(df_plot_time, df_plot['temp.'], marker='', label='Temperature (°C)')

# Format the x-axis to show HH:MM:SS
ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))


plt.xlabel('Time')
plt.ylabel('Sensor Readings')
plt.title('Sensor Readings vs Time')


plt.legend()
plt.xticks(rotation=45)   # rotate lable for ease of reading 
plt.tight_layout()
plt.show()
